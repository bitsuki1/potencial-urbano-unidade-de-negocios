# Gera ESTEIRA-PENDENCIAS-PU-<data>.xlsx a partir do BACKLOG/mapa. Requer: pip install openpyxl. Rode na raiz do repo.
# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Esteira de Pendencias"

# ---- estilos ----
F_TITULO = Font(bold=True, size=14, color="FFFFFF")
FILL_TITULO = PatternFill("solid", fgColor="1F4E78")
F_HEAD = Font(bold=True, color="FFFFFF", size=10)
FILL_HEAD = PatternFill("solid", fgColor="2E75B6")
FILL_ETAPA = PatternFill("solid", fgColor="D9E1F2")
F_ETAPA = Font(bold=True, size=10, color="1F4E78")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
# status
FILL_FEITO = PatternFill("solid", fgColor="C6EFCE")   # verde
FILL_PARCIAL = PatternFill("solid", fgColor="FFEB9C")  # amarelo
FILL_ABERTO = PatternFill("solid", fgColor="FFC7CE")   # vermelho claro
FILL_RESOLV = PatternFill("solid", fgColor="C6EFCE")
# de quem
FILL_SEU = PatternFill("solid", fgColor="DDEBF7")      # azul (projeto)
FILL_MEU = PatternFill("solid", fgColor="FCE4D6")      # laranja (MOU)

cols = ["Etapa", "#", "Pendencia", "Artefato", "O que falta (prova/DoD)", "Prioridade", "Status", "De quem", "Bloqueio"]
widths = [22, 7, 40, 14, 46, 12, 12, 11, 26]

# ---- titulo ----
ws.merge_cells("A1:I1")
ws["A1"] = "POTENCIAL URBANO — Esteira de Pendencias (IPTU/TDC)"
ws["A1"].font = F_TITULO; ws["A1"].fill = FILL_TITULO; ws["A1"].alignment = CENTER
ws.merge_cells("A2:I2")
ws["A2"] = "Gerado 2026-06-21 (chapeu PU). SSOT vivo = BACKLOG.md + MANIFESTO.json. Prioridade: vermelho=produto, azul=corpus/RAG, amarelo=higiene, branco=qualidade."
ws["A2"].font = Font(italic=True, size=9, color="595959"); ws["A2"].alignment = Alignment(wrap_text=True)
ws.row_dimensions[2].height = 26

HEAD_ROW = 4
for j, c in enumerate(cols, 1):
    cell = ws.cell(HEAD_ROW, j, c)
    cell.font = F_HEAD; cell.fill = FILL_HEAD; cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center"); cell.border = BORDER
for j, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(j)].width = w

# ---- dados: (etapa, #, pendencia, artefato, oque_falta, prioridade, status, dequem, bloqueio) ----
P = "🟥 produto"; B = "🟦 corpus"; H = "🟨 higiene"; Q = "⬜ qualidade"; DASH="—"
rows = [
 ("FASE 0 — Limpar/Organizar Drive", DASH, "Limpar duplicatas no Drive (~16-20 GB)", "Geo/Cadastral", "Rodar a dedup (lixeira recuperavel ~30d).", H, "ABERTO", "MEU", "voce retomou via projeto do escritorio (Dedup-V2)"),
 ("FASE 0 — Limpar/Organizar Drive", "B-8", "AUD-02: conferir IDs canonicos ANTES de qualquer DELETE", "Meta", "Cada ID 'manter' conferido por get_file_metadata antes do delete (risco ~3 GB).", H, "ABERTO", "MEU+SEU", "lane do Drive (orquestrador deposita)"),
 ("FASE 0 — Limpar/Organizar Drive", DASH, "Arrumacao fisica do Drive (992 itens)", "Meta", DASH, DASH, "FEITO", "—", "historico (Fase 0)"),

 ("ETAPA 1 — Fundacao canonica", "B-9", "Descer do Drive o verbatim TDC + Q14 + Quadro 3", "Lei/Tabela", "Maestro relaya a lane; lane puxa p/ _entrada/. GO do MOU dado.", P, "ABERTO", "MEU+SEU", "GARGALO UNICO do produto"),
 ("ETAPA 1 — Fundacao canonica", "B-9/P2", "Re-ingerir corpus TDC verbatim (PDE 16.050/2014)", "Lei", "fatiar+indexar; os 3 evals TDC viram gate de aceite.", P, "ABERTO", "SEU", "Drive (B-9)"),
 ("ETAPA 1 — Fundacao canonica", "B-4", "Re-ingerir as 14 leis municipais em verbatim", "Lei", "cada .md verbatim + .json + indexada (promover_entrada).", B, "ABERTO", "SEU", "Drive (cru nao-local)"),
 ("ETAPA 1 — Fundacao canonica", DASH, "Taguear tudo (taxonomia facetada, secao 2)", "Meta", "tags por conteudo, nao por titulo.", H, "PARCIAL", "SEU", "corpus verbatim"),
 ("ETAPA 1 — Fundacao canonica", "B-15", "Forjar o CODEX-MESTRE IPTU/TDC", "Tese/Meta", "consolidar matrizes -> escrutinar doc-a-doc (Tese/Antitese/Vacina). Nome ja decidido.", B, "ABERTO", "SEU", "corpus verbatim (sob D-13: voce escrutina)"),
 ("ETAPA 1 — Fundacao canonica", "B-10", "Auditar o MERITO JURIDICO das teses", "Tese", "revisao tema-a-tema com citacao verificada + vacinas.", Q, "ABERTO", "SEU", "tempo/expertise"),

 ("RAG — fatiar/indexar/consultar", DASH, "Tubo RAG fim-a-fim (citacao obrigatoria 1.7)", "Lei", DASH, DASH, "FEITO", "—", "entregue (13 leis indexadas)"),
 ("RAG — fatiar/indexar/consultar", "B-11", "Chunker: vigencia POR CHUNK (280 rotulos duplicados)", "Lei", "redacao revogada nao volta como vigente; eval de redacao-compilada.", B, "PARCIAL", "SEU", "nenhum (local) — atacavel ja"),
 ("RAG — fatiar/indexar/consultar", "B-5", "Camada semantica (embeddings) + filtro por tema", "Lei", "'direito de construir' deixa de casar a lei errada; eval armadilha lexical.", B, "ABERTO", "SEU", "nenhum (local)"),
 ("RAG — fatiar/indexar/consultar", "B-6", "Grafo de remissoes / vigencia por remissao", "Lei", "data por remissao entre artigos vira positivo.", B, "ABERTO", "SEU", "nenhum (local)"),
 ("RAG — fatiar/indexar/consultar", "B-7", "Vigencia municipal DATADA + campo verbatim_integral", "Lei/Meta", "15 municipais com vigencia.inicio/fim na cadeia.", B, "ABERTO", "SEU", "nenhum (local)"),

 ("ESTRUTURA DE CONSUMO — engine+banco", DASH, "Engine TDC em codigo (oodc.py determinístico)", "Formula", DASH, DASH, "FEITO", "—", "entregue (auto-testado no CI)"),
 ("ESTRUTURA DE CONSUMO — engine+banco", "B-1", "Ingerir tabelas Q14 + Quadro 3 -> tabelas/", "Tabela", "csv com proveniencia; engine sobre V e CA reais de >=1 imovel.", P, "ABERTO", "SEU", "Drive (B-9)"),
 ("ESTRUTURA DE CONSUMO — engine+banco", "B-3", "Completar tabelas Fs/Fp no engine", "Formula", "FATOR_SOCIAL/PLANEJAMENTO completos + autoteste; gate verde.", P, "ABERTO", "SEU", "B-1"),
 ("ESTRUTURA DE CONSUMO — engine+banco", "B-12", "Acabamento do engine (decimal/FATAL/citacao por artigo)", "Formula", "OODC que estoura o tipo LEVANTA; cita o artigo, nao a lei inteira.", P, "PARCIAL", "SEU", "parcial B-1"),
 ("ESTRUTURA DE CONSUMO — engine+banco", "P5", "Criar schemas Supabase (artefatos + geo + rag)", "Banco", "DDL apos organizacao aprovada (RO-23).", B, "ABERTO", "SEU", "organizacao aprovada"),

 ("FASE 2 — Identificacao de imoveis", "B-2", "1o JOIN do produto (IPTU2026 x LOTES x Q14 x zoneamento)", "Cadastral/Geo", "saida com N imoveis reais {SQL, valor, oportunidade, dono}.", P, "ABERTO", "SEU", "B-1 + dados pesados (Drive->Supabase)"),
 ("FASE 2 — Identificacao de imoveis", "E2", "Indexar GeoSampa/zoneamento/MASTER_PARAMETROS", "Geo/Tabela", "camadas lote a lote; PGV/valor venal por SQL.", B, "ABERTO", "SEU", "Drive"),
 ("FASE 2 — Identificacao de imoveis", "E3", "Base de proprietarios (IPTU2026, ITBI 2006-2024, socios)", "Cadastral", "indexar CSVs pesados; cruzar dono.", B, "ABERTO", "SEU", "Drive->Supabase"),

 ("FASE 3 — Validacao por casos reais", DASH, "Ground-truth + engine campo a campo", "Tese/Formula", "formula que nao bate com caso real NAO passa.", B, "ABERTO", "SEU", "corpus + engine com dado"),

 ("PRODUTO (E5) — o entregavel", DASH, "Lista de alvos por imovel (oportunidade+valor+dono)", "Produto", "dossie de prospeccao por imovel.", P, "ABERTO", "SEU", "toda a cadeia acima"),

 ("DECISOES / ACOES DO MOU", "M-41", "Banco PU: fechar a porta (remover 'public' dos Exposed schemas)", "Banco", "1 clique no painel Supabase (fecha o aviso de seguranca).", H, "ABERTO", "MEU", "—"),
 ("DECISOES / ACOES DO MOU", "M-24", "Direcao do corpus (fechar E1 so x abrir E2/E3)", "Meta", "DECIDIDO: serial, um de cada vez (D-13).", DASH, "RESOLVIDO", "MEU", "—"),
 ("DECISOES / ACOES DO MOU", "B-16", "Escopo geografico", "Meta", "SP-capital agora; expande depois.", DASH, "RESOLVIDO", "MEU", "—"),
 ("DECISOES / ACOES DO MOU", "B-15", "Nome do Codex Mestre", "Meta", "CODEX-MESTRE IPTU/TDC.", DASH, "RESOLVIDO", "MEU", "—"),
 ("DECISOES / ACOES DO MOU", DASH, "Itens fora de foco (ISS / previdenciario)", "Lei", "relocados p/ jurisprudencia/_correlatos/ (preservados).", DASH, "RESOLVIDO", "MEU", "—"),
]

r = HEAD_ROW + 1
last_etapa = None
for (etapa, num, pend, art, falta, prio, status, dequem, bloq) in rows:
    if etapa != last_etapa:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        c = ws.cell(r, 1, etapa)
        c.fill = FILL_ETAPA; c.font = F_ETAPA; c.alignment = Alignment(vertical="center")
        c.border = BORDER
        ws.row_dimensions[r].height = 18
        r += 1
        last_etapa = etapa
    vals = ["", num, pend, art, falta, prio, status, dequem, bloq]
    for j, v in enumerate(vals, 1):
        cell = ws.cell(r, j, v)
        cell.alignment = WRAP; cell.border = BORDER; cell.font = Font(size=9)
        if j == 6: cell.alignment = CENTER
        if j == 7:
            cell.alignment = CENTER
            cell.fill = {"FEITO":FILL_FEITO,"RESOLVIDO":FILL_RESOLV,"PARCIAL":FILL_PARCIAL,"ABERTO":FILL_ABERTO}.get(status)
        if j == 8:
            cell.alignment = CENTER
            cell.fill = FILL_SEU if "SEU" in dequem and "MEU" not in dequem else (FILL_MEU if dequem.startswith("MEU") else FILL_SEU)
            if dequem == "MEU+SEU": cell.fill = FILL_MEU
    ws.row_dimensions[r].height = 30
    r += 1

ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A{HEAD_ROW}:I{r-1}"

# ---- aba Legenda ----
lg = wb.create_sheet("Legenda")
leg = [
 ("LEGENDA", ""),
 ("De quem", ""),
 ("SEU (azul)", "o orquestrador do Potencial Urbano resolve no repo, sem depender de voce"),
 ("MEU (laranja)", "acao fisica ou decisao do MOU (so voce fecha)"),
 ("MEU+SEU", "o projeto prepara/deposita; voce (ou a lane do Drive) executa"),
 ("", ""),
 ("Status", ""),
 ("FEITO/RESOLVIDO (verde)", "concluido e provado"),
 ("PARCIAL (amarelo)", "comecado, falta a parte marcada na coluna 'o que falta'"),
 ("ABERTO (vermelho)", "ainda nao iniciado / aguardando"),
 ("", ""),
 ("Prioridade", ""),
 ("vermelho = produto/valor", "o que da dinheiro (lista de alvos)"),
 ("azul = corpus/RAG", "leis, fatiamento, busca"),
 ("amarelo = higiene/governanca", "Drive, seguranca, organizacao"),
 ("branco = qualidade", "melhoria, nao bloqueia"),
 ("", ""),
 ("Nota", "O caminho do dinheiro hoje passa TODO pelo B-9 (descer os dados do Drive). GO ja dado. SSOT vivo = BACKLOG.md."),
]
lg.column_dimensions["A"].width = 30; lg.column_dimensions["B"].width = 70
for i,(a,b) in enumerate(leg,1):
    lg.cell(i,1,a).font = Font(bold=(b=="" and a!=""), size=11 if a in ("LEGENDA",) else 10)
    lg.cell(i,2,b).alignment = Alignment(wrap_text=True)
lg["A1"].font = Font(bold=True, size=13, color="1F4E78")

out = "ESTEIRA-PENDENCIAS-PU-2026-06-21.xlsx"
wb.save(out)
print("salvo:", out, "| linhas de dados:", r-HEAD_ROW-1)
