# -*- coding: utf-8 -*-
import csv, sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RAIZ = Path(__file__).resolve().parents[2]
SRC = RAIZ / "zepec" / "ferramenta" / "lista_prospeccao.csv"
OUT_DEFAULT = RAIZ / "zepec" / "ferramenta" / "Potencial-Urbano_Lista-Prospeccao-ZEPEC.xlsx"

COLS = [
    ("segmento",              "Segmento (estágio·dono)"),
    ("estado_venda",          "Estágio de venda"),
    ("nome_bem",              "Bem tombado"),
    ("endereco_mestre",       "Endereço"),
    ("distrito",              "Distrito"),
    ("proprietario",          "Proprietário"),
    ("tipo_zepec",            "Tipo ZEPEC"),
    ("esfera",                "Esferas (quem tombou)"),
    ("m2_ja_transferido",     "m² já transferido"),
    ("preco_legal_ref_brl",   "Preço legal ref. (R$) — Art. 128"),
    ("status_fundurb",        "Status FUNDURB do cedente"),
    ("intercorrencia_fundurb","Intercorrência FUNDURB"),
    ("data_ref",              "Data ref."),
    ("sql_mestre",            "SQL (chave do imóvel)"),
]

AZUL_ESCURO = "1F4E78"; AZUL = "2E75B6"; CINZA = "F2F2F2"
COR_ESTAGIO = {
    "INTACTO":    "C6EFCE",
    "TEM_SALDO":  "FFEB9C",
    "SO_ELEGIVEL":"BDD7EE",
    "INCERTO":    "F2F2F2",
}


def gerar_xlsx(out_path=None):
    out = Path(out_path) if out_path else OUT_DEFAULT
    rows = list(csv.DictReader(open(SRC, encoding="utf-8")))

    f_titulo = Font(bold=True, size=15, color="FFFFFF")
    fill_titulo = PatternFill("solid", fgColor=AZUL_ESCURO)
    f_head = Font(bold=True, color="FFFFFF", size=10)
    fill_head = PatternFill("solid", fgColor=AZUL)
    fill_zebra = PatternFill("solid", fgColor=CINZA)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Lista de Prospecção"

    ncol = len(COLS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    c = ws.cell(row=1, column=1, value="POTENCIAL URBANO — Lista de Prospecção ZEPEC (cedentes de TDC)")
    c.font = f_titulo; c.fill = fill_titulo; c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    sc = ws.cell(row=2, column=1,
                 value=f"{len(rows)} imóveis prontos para abordar  ·  só fato, sem juízo de valor (a priorização é sua)  ·  base completa: 6.131 imóveis")
    sc.font = Font(italic=True, size=9, color="595959"); sc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    head_row = 3
    for j, (_, titulo) in enumerate(COLS, start=1):
        cell = ws.cell(row=head_row, column=j, value=titulo)
        cell.font = f_head; cell.fill = fill_head; cell.alignment = center; cell.border = border
    ws.row_dimensions[head_row].height = 30

    for i, r in enumerate(rows):
        rownum = head_row + 1 + i
        for j, (key, _) in enumerate(COLS, start=1):
            val = (r.get(key, "") or "").strip()
            if val == "—": val = ""
            if key == "m2_ja_transferido" and val:
                try: val = round(float(val), 2)
                except ValueError: pass
            cell = ws.cell(row=rownum, column=j, value=val)
            if key == "preco_legal_ref_brl" and val:
                try:
                    cell.value = float(val); cell.number_format = 'R$ #,##0.00'
                except ValueError:
                    pass
            cell.alignment = center if key in ("estado_venda","tipo_zepec","esfera","distrito","m2_ja_transferido","preco_legal_ref_brl","data_ref","sql_mestre","status_fundurb") else left
            cell.border = border
            cell.font = Font(size=9)
            if i % 2 == 1:
                cell.fill = fill_zebra
        est = (r.get("estado_venda","") or "").strip()
        if est in COR_ESTAGIO:
            cc = ws.cell(row=rownum, column=2)
            cc.fill = PatternFill("solid", fgColor=COR_ESTAGIO[est])
            cc.font = Font(size=9, bold=True)

    larg = {"Segmento (estágio·dono)":20,"Estágio de venda":15,"Bem tombado":34,"Endereço":30,"Distrito":16,
            "Proprietário":34,"Tipo ZEPEC":12,"Esferas (quem tombou)":22,"m² já transferido":14,
            "Preço legal ref. (R$) — Art. 128":22,
            "Status FUNDURB do cedente":22,"Intercorrência FUNDURB":20,"Data ref.":12,"SQL (chave do imóvel)":18}
    for j, (_, titulo) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = larg.get(titulo, 16)

    ws.freeze_panes = f"A{head_row+1}"
    ws.auto_filter.ref = f"A{head_row}:{get_column_letter(ncol)}{head_row+len(rows)}"

    wl = wb.create_sheet("Legenda")
    wl.column_dimensions["A"].width = 22; wl.column_dimensions["B"].width = 70
    tc = wl.cell(row=1, column=1, value="Como ler a planilha"); tc.font = Font(bold=True, size=13, color="FFFFFF"); tc.fill = fill_titulo
    wl.merge_cells("A1:B1")
    legenda = [
        ("Estágio de venda", ""),
        ("INTACTO", "Declarou potencial e NUNCA vendeu — potencial cheio. Abordar."),
        ("TEM_SALDO", "Vendeu parte, ainda resta potencial. Abordar (calcular o que resta)."),
        ("SO_ELEGIVEL", "Tombado que ainda NÃO declarou — pode entrar (precisa declarar antes)."),
        ("", ""),
        ("Esferas", "Quem tombou o bem: municipal (CONPRESP) / estadual (CONDEPHAAT) / federal (IPHAN). Mais esferas = mais órgãos na negociação."),
        ("m² já transferido", "Quanto de potencial o imóvel já vendeu (0 = intacto)."),
        ("SQL", "'CPF do imóvel' (setor-quadra-lote) — a chave que liga tudo."),
        ("Status FUNDURB do cedente", "Status do processo pecuniário FUNDURB do próprio cedente. Hoje INDETERMINADO até confirmar a semântica com a Prefeitura (SMUL)."),
        ("Proprietário", "Cobertura PARCIAL hoje — sobe em escala quando o IPTU/ITBI subir ao Supabase."),
        ("Preço legal ref. (R$)", "Referência legal do Art. 128 (PDE): (potencial × valor do terreno) ÷ 4 — o piso regulatório, que INDEPENDE do comprador (o Cr dele cancela na conta). A MARGEM é sua. A memória de cálculo completa fica no DOSSIÊ de cada imóvel (gerar_dossie.py)."),
        ("Princípio", "Só fato, sem 'vale/melhor/pior'. A priorização comercial é do time. Preço nasce do engine, nunca é inventado."),
    ]
    for i, (a, b) in enumerate(legenda, start=2):
        ca = wl.cell(row=i, column=1, value=a); ca.font = Font(bold=bool(a and not b), size=10); ca.alignment = Alignment(vertical="top", wrap_text=True)
        cb = wl.cell(row=i, column=2, value=b); cb.font = Font(size=10); cb.alignment = Alignment(vertical="top", wrap_text=True)
        if a in COR_ESTAGIO:
            ca.fill = PatternFill("solid", fgColor=COR_ESTAGIO[a])

    # RESUMO (funil) — a manchete: primeira aba, lida de funil.csv (gerada por funil.py).
    FUNIL = RAIZ / "zepec" / "ferramenta" / "funil.csv"
    if FUNIL.exists():
        wr = wb.create_sheet("Resumo (funil)", 0)
        for col, w in (("A", 46), ("B", 12), ("C", 10), ("D", 62)):
            wr.column_dimensions[col].width = w
        tr = wr.cell(row=1, column=1, value="POTENCIAL URBANO — Resumo do funil (cedentes de TDC)")
        tr.font = Font(bold=True, size=14, color="FFFFFF"); tr.fill = fill_titulo
        wr.merge_cells("A1:D1"); wr.row_dimensions[1].height = 24
        frows = list(csv.DictReader(open(FUNIL, encoding="utf-8")))
        rr = 3
        for bloco, titulo in (("completude", "Completude do dado (só decresce)"),
                              ("corte_acao", "Cortes de ação (a fila de abordagem)")):
            bc = wr.cell(row=rr, column=1, value=titulo)
            bc.font = Font(bold=True, size=11, color="FFFFFF"); bc.fill = fill_head
            wr.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=4); rr += 1
            for fr in [x for x in frows if x["bloco"] == bloco]:
                wr.cell(row=rr, column=1, value=fr["estagio"]).font = Font(size=10)
                qc = wr.cell(row=rr, column=2, value=int(fr["quantidade"]))
                qc.font = Font(size=10, bold=True); qc.alignment = Alignment(horizontal="center")
                wr.cell(row=rr, column=3, value=fr["do_total_pct"]).alignment = Alignment(horizontal="center")
                nc = wr.cell(row=rr, column=4, value=fr["nota"]); nc.font = Font(size=9, color="595959")
                nc.alignment = Alignment(wrap_text=True, vertical="center")
                rr += 1
            rr += 1

    wb.save(out)
    print(f"gerado: {out} | imóveis: {len(rows)}")


if __name__ == "__main__":
    dest = sys.argv[1] if len(sys.argv) > 1 else None
    gerar_xlsx(dest)
