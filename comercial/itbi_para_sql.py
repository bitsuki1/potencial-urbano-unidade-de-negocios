#!/usr/bin/env python3
"""
itbi_para_sql.py — normaliza as GUIAS DE ITBI PAGAS (Fazenda SP) em um índice por SQL: a transação MAIS
RECENTE de cada imóvel (valor + data + natureza). Base do "preço real" e do "frescor do dono" na lista.

Fonte PRIMÁRIA (1.3): guias oficiais da SF-SP (SQL·valor de transação·data·natureza·matrícula·cartório·uso).
O ITBI NÃO traz nome (Fazenda anonimiza) — este índice é só valor+data por SQL. As guias vêm em xlsx (série
2006→2024 + snapshot 2026). Aceita xlsx (openpyxl) e csv (fixture). Colunas casadas por NOME (robusto à ordem).

Saída: zepec/oficial/itbi_por_sql.csv → sql_mestre, valor_transacao, data_transacao, natureza, fonte_ano.

Uso:
  python3 comercial/itbi_para_sql.py --entrada guias_2023.xlsx guias_2024.xlsx ...
  python3 comercial/itbi_para_sql.py --autoteste
"""
import csv
import re
import sys
import argparse
import unicodedata
from pathlib import Path

RAIZ = Path(__file__).resolve().parent.parent
SAIDA = RAIZ / "zepec" / "oficial" / "itbi_por_sql.csv"
FIX = RAIZ / "evals" / "ground-truth" / "comercial-fixture" / "itbi_guia_amostra.csv"


def _norm(s):
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", s).strip().upper()


def _sql10(s):
    d = re.sub(r"\D", "", str(s or ""))
    return d[:10] if len(d) >= 10 else ""


def _data_iso(s):
    s = str(s or "").strip()
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return m.group(0)
    m = re.search(r"(\d{2})/(\d{2})/(\d{4})", s)      # dd/mm/aaaa → aaaa-mm-dd
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    return ""


def _num(s):
    s = str(s or "").strip()
    if not s:
        return None
    s = s.replace("R$", "").strip()
    if "," in s and "." in s:      # 1.234.567,89
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:                  # 1234,89
        s = s.replace(",", ".")
    try:
        return float(re.sub(r"[^\d.\-]", "", s))
    except Exception:
        return None


def _linhas_xlsx(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        yield [("" if c is None else c) for c in row]
    wb.close()


def _linhas_csv(path):
    with open(path, encoding="utf-8", errors="replace") as f:
        amostra = f.readline()
        delim = ";" if amostra.count(";") >= amostra.count(",") else ","
        f.seek(0)
        for row in csv.reader(f, delimiter=delim):
            yield row


def indexar(fontes, ano_por_fonte=None):
    """fontes = lista de (nome, iterável-de-linhas). Devolve dict sql → melhor transação (a mais recente)."""
    melhor = {}
    for nome, linhas in fontes:
        it = iter(linhas)
        try:
            cab = [_norm(c) for c in next(it)]
        except StopIteration:
            continue
        def idx(*alvos):
            for i, c in enumerate(cab):
                for a in alvos:
                    if _norm(a) in c:
                        return i
            return -1
        i_sql = idx("CADASTRO (SQL)", "N DO CADASTRO", "NUMERO DO CADASTRO", "SQL")
        i_val = idx("VALOR DE TRANSACAO", "VALOR DA TRANSACAO", "VALOR DE TRANSACAO (DECLARADO")
        i_dat = idx("DATA DE TRANSACAO", "DATA DA TRANSACAO")
        i_nat = idx("NATUREZA DE TRANSACAO", "NATUREZA DA TRANSACAO", "NATUREZA")
        if i_sql < 0 or i_dat < 0:
            continue
        for row in it:
            if not row:
                continue
            sql = _sql10(row[i_sql] if i_sql < len(row) else "")
            data = _data_iso(row[i_dat] if 0 <= i_dat < len(row) else "")
            if not sql or not data:
                continue
            val = _num(row[i_val]) if 0 <= i_val < len(row) else None
            nat = str(row[i_nat]).strip() if 0 <= i_nat < len(row) else ""
            cur = melhor.get(sql)
            if cur is None or data > cur["data_transacao"]:
                melhor[sql] = {"sql_mestre": sql, "valor_transacao": ("" if val is None else f"{val:.2f}"),
                               "data_transacao": data, "natureza": nat, "fonte_ano": nome}
    return melhor


def _autoteste():
    melhor = indexar([("fixture", _linhas_csv(FIX))])
    # duas transações do mesmo SQL → fica a MAIS RECENTE (2023, não 2015)
    assert melhor["0200670033"]["data_transacao"] == "2023-05-10", melhor["0200670033"]
    assert melhor["0200670033"]["valor_transacao"] == "4804259.04"
    # valor com formato brasileiro (1.234.567,89) parseado
    assert melhor["0100010001"]["valor_transacao"] == "767498.59", melhor["0100010001"]
    return len(melhor)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--entrada", nargs="*", default=[])
    ap.add_argument("--autoteste", action="store_true")
    args = ap.parse_args()
    if args.autoteste:
        n = _autoteste()
        print(f"AUTO-TESTE itbi_para_sql: OK — {n} SQLs; mais-recente-vence e parsing R$ conferidos.")
        return 0
    if not args.entrada:
        print("uso: --entrada <guia.xlsx|csv> [...]  ou  --autoteste", file=sys.stderr); return 2
    fontes = []
    for p in args.entrada:
        linhas = _linhas_xlsx(p) if str(p).lower().endswith((".xlsx", ".xlsm")) else _linhas_csv(p)
        fontes.append((Path(p).name, linhas))
    melhor = indexar(fontes)
    SAIDA.parent.mkdir(parents=True, exist_ok=True)
    with open(SAIDA, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["sql_mestre", "valor_transacao", "data_transacao", "natureza", "fonte_ano"])
        w.writeheader(); w.writerows(melhor.values())
    print(f"OK: {len(melhor)} SQLs com transação → zepec/oficial/itbi_por_sql.csv")
    return 0


if __name__ == "__main__":
    sys.exit(main())
